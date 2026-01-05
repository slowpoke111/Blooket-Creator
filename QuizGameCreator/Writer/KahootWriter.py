from QuizGameCreator.Writer.QuestionBank import QuestionBank
from QuizGameCreator.Writer.Writer import Writer
from QuizGameCreator.Writer.Question import Question
from typing import List

class KahootWriter(Writer):
    skipAI = False
    outputFormat = "xlsx"
    
    def __init__(self, template_path:str = "QuizGameCreator/Writer/Templates/Kahoot.csv") -> None:
        self.template_path = template_path
        self.questions: List[Question] = []
    
    def add_question(self, question: Question) -> None:
        self.questions.append(question)
        
    def add_questions(self, questions: List[Question]) -> None:
        self.questions.extend(questions)
    
    def _question_to_row(self, question: Question, question_number: int = -1) -> List[str]:
        answers = question.answers
        correct_answers = ",".join(str(i) for i in question.correct_answers)
        return [
            str(question_number),
            question.question_text,
            answers[0],
            answers[1],
            answers[2],
            answers[3],
            str(question.time_limit),
            correct_answers
        ]
    def write(self, output_path: str) -> None:
        from openpyxl import Workbook, load_workbook
        import csv
        from openpyxl.utils import get_column_letter

        with open(self.template_path, 'r', encoding='utf-8') as template_file:
            reader = csv.reader(template_file)
            template_rows = list(reader)

        wb = Workbook()
        ws = wb.active

        for row in template_rows:
            ws.append(row) #type:ignore

        start_row = len(template_rows) + 1

        for i, question in enumerate(self.questions, start=1):
            row = self._question_to_row(question, question_number=i)
            ws.append(row) #type:ignore

        for col in range(1, ws.max_column + 1): #type:ignore
            ws.column_dimensions[get_column_letter(col)].width = 30 #type:ignore

        wb.save(output_path)
    
    def clear(self) -> bool:
        self.questions.clear()
        return True
        